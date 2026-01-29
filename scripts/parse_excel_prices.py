#!/usr/bin/env python3
"""
Script pour parser le fichier Excel ELEMENTS BOOK DCM et extraire tous les prix
"""

import openpyxl
import json
import sys

def parse_excel_prices(excel_path):
    """Parse le fichier Excel et extrait tous les prix et marges"""
    
    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        print(f"Fichier Excel chargé : {excel_path}")
        print(f"Feuilles disponibles : {workbook.sheetnames}")
        
        all_data = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"\n=== Analyse de la feuille : {sheet_name} ===")
            print(f"Dimensions : {sheet.max_row} lignes x {sheet.max_column} colonnes")
            
            # Afficher les 20 premières lignes pour comprendre la structure
            print("\nPremières lignes :")
            for i, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), 1):
                # Filtrer les lignes vides
                if any(cell is not None for cell in row):
                    print(f"Ligne {i}: {row}")
            
            all_data[sheet_name] = {
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'sample_data': []
            }
            
            # Stocker un échantillon de données
            for row in sheet.iter_rows(max_row=50, values_only=True):
                if any(cell is not None for cell in row):
                    all_data[sheet_name]['sample_data'].append(row)
        
        return all_data
        
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier Excel : {e}")
        return None

if __name__ == "__main__":
    excel_path = "/Users/macbook/Documents/SABC/lions_book/assets/ELEMENTS BOOK DCM.xlsx"
    
    print("Analyse du fichier Excel ELEMENTS BOOK DCM...")
    print("=" * 80)
    
    data = parse_excel_prices(excel_path)
    
    if data:
        # Sauvegarder les données extraites
        output_path = "/tmp/excel_prices_analysis.json"
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n\nAnalyse sauvegardée dans : {output_path}")
        print("\nRésumé :")
        for sheet_name, sheet_data in data.items():
            print(f"  - {sheet_name}: {sheet_data['max_row']} lignes, {sheet_data['max_column']} colonnes")
    else:
        print("Échec de l'analyse du fichier Excel")
        sys.exit(1)
