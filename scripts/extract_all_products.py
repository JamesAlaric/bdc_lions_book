#!/usr/bin/env python3
"""
Script pour extraire tous les produits du fichier Excel ELEMENTS BOOK DCM
et générer les fichiers YAML pour chaque segment
"""

import openpyxl
import json
from pathlib import Path

# Mapping codes marques vers noms
BRAND_MAPPING = {
    # Bières
    'EXP': '33 Export',
    'CAS': 'Castel Beer',
    'MUT': 'Mützig',
    'MNY': 'Manyan',
    'ISE': 'Isenbeck',
    'BFT': 'Beaufort Lager',
    'BFL': 'Beaufort Light',
    'CHC': 'Chill Citron',
    'DOP': 'Doppel Munich',
    'DOL': 'Doppel Lager',
    'CMS': 'Castle Milk Stout',
    'HEI': 'Heineken',
    # Alcools Mix
    'BWC': 'Booster Whisky Cola',
    'BGT': 'Booster Gin Tonic',
    'SMP': 'Smirnoff Ice Pineapple',
    'SMB': 'Smirnoff Ice Black',
    'ORI': 'Orijin',
    # Boissons Gazeuses
    'TGI': 'TOP Grenadine',
    'TPA': 'TOP Ananas',
    'TPG': 'TOP Grenadine',
    'TPO': 'TOP Orange',
    'TPP': 'TOP Pamplemousse',
    'TPT': 'TOP Tropical',
    'WCO': 'World Cola',
    'YOU': 'Youzou',
    'ORG': 'Orangina',
    'DJC': "D'jino",
    'VIM': 'Vimto',
    'SOD': 'Soda',
    # Eaux
    'VIT': 'Vitale',
    'TCI': 'Tangui Citron',
    'AQB': 'Aquabelle'
}

def extract_products_from_sheet(sheet, segment_name):
    """Extrait les produits d'une feuille Excel"""
    
    products = []
    current_brand = None
    
    # Trouver la ligne d'en-tête (contient "Code article")
    header_row = None
    for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row and 'Code article' in str(row):
            header_row = i
            break
    
    if not header_row:
        print(f"  ⚠️  Ligne d'en-tête non trouvée dans {segment_name}")
        return products
    
    # Extraire les données à partir de la ligne suivant l'en-tête
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not any(row):
            continue
        
        # Structure attendue des colonnes
        segment = row[0] if len(row) > 0 else None
        marque = row[1] if len(row) > 1 else None
        code = row[2] if len(row) > 2 else None
        designation = row[3] if len(row) > 3 else None
        format_val = row[4] if len(row) > 4 else None
        unite = row[5] if len(row) > 5 else None
        emballage = row[6] if len(row) > 6 else None
        consigne = row[7] if len(row) > 7 else None
        marche = row[8] if len(row) > 8 else None
        
        # Prix Marketing → Distributeur
        prix_achat_mkt = row[19] if len(row) > 19 else None
        remise_mkt = row[20] if len(row) > 20 else None
        frais_chr = row[21] if len(row) > 21 else None
        marge_mkt = row[22] if len(row) > 22 else None
        taux_mkt = row[23] if len(row) > 23 else None
        
        # Prix Distributeur → Détaillant
        prix_achat_dist = row[25] if len(row) > 25 else None
        ristourne = row[26] if len(row) > 26 else None
        frais_elv = row[27] if len(row) > 27 else None
        marge_dist = row[28] if len(row) > 28 else None
        taux_dist = row[29] if len(row) > 29 else None
        
        # Prix Consommateur
        prix_unitaire = row[31] if len(row) > 31 else None
        prix_casier = row[32] if len(row) > 32 else None
        
        # Mettre à jour la marque courante si présente
        if marque:
            current_brand = marque
        
        # Ignorer les lignes sans code produit
        if not code:
            continue
        
        # Créer l'objet produit
        product = {
            'code': code,
            'brand': current_brand or 'Unknown',
            'designation': designation,
            'format': format_val,
            'unit': unite,
            'packaging': emballage,
            'consigne': int(consigne) if consigne else 0,
            'market': marche,
            'pricing': {}
        }
        
        # Ajouter les prix si disponibles
        if prix_achat_mkt is not None:
            product['pricing']['marketing_to_distributor'] = {
                'prix_achat': round(float(prix_achat_mkt), 2) if prix_achat_mkt else None,
                'remise': round(float(remise_mkt), 2) if remise_mkt else None,
                'frais_chr': round(float(frais_chr), 2) if frais_chr else None,
                'marge': round(float(marge_mkt), 2) if marge_mkt else None,
                'taux_marge': f"{round(float(taux_mkt) * 100, 1)}%" if taux_mkt else None
            }
        
        if prix_achat_dist is not None:
            product['pricing']['distributor_to_retailer'] = {
                'prix_achat': round(float(prix_achat_dist), 2) if prix_achat_dist else None,
                'ristourne': round(float(ristourne), 2) if ristourne else None,
                'frais_elv': round(float(frais_elv), 2) if frais_elv else None,
                'marge': round(float(marge_dist), 2) if marge_dist else None,
                'taux_marge': f"{round(float(taux_dist) * 100, 1)}%" if taux_dist else None
            }
        
        if prix_unitaire is not None:
            product['pricing']['consumer_price'] = {
                'prix_unitaire': int(prix_unitaire) if prix_unitaire else None,
                'prix_casier': int(prix_casier) if prix_casier else None
            }
        
        products.append(product)
    
    return products

def group_by_brand(products):
    """Groupe les produits par marque"""
    brands = {}
    
    for product in products:
        brand_name = product['brand']
        if brand_name not in brands:
            brands[brand_name] = []
        brands[brand_name].append(product)
    
    return brands

def main():
    excel_path = "/Users/macbook/Documents/SABC/lions_book/assets/ELEMENTS BOOK DCM.xlsx"
    output_dir = Path("/Users/macbook/Documents/SABC/lions_book/data/static/catalog")
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print("Extraction des produits du fichier Excel...")
    print("=" * 80)
    
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    
    all_segments = {}
    
    # Extraire chaque segment
    segments_config = {
        'BIERES': 'bieres',
        'AM': 'alcools-mix',
        'BG': 'boissons-gazeuses',
        'EAUX': 'eaux'
    }
    
    for sheet_name, segment_id in segments_config.items():
        if sheet_name in workbook.sheetnames:
            print(f"\n📦 Extraction segment : {sheet_name}")
            sheet = workbook[sheet_name]
            products = extract_products_from_sheet(sheet, sheet_name)
            brands = group_by_brand(products)
            
            print(f"  ✅ {len(products)} produits extraits")
            print(f"  ✅ {len(brands)} marques identifiées")
            
            all_segments[segment_id] = {
                'products': products,
                'brands': brands
            }
            
            # Afficher résumé par marque
            for brand_name, brand_products in sorted(brands.items()):
                print(f"     - {brand_name}: {len(brand_products)} produits")
    
    # Sauvegarder en JSON
    output_json = output_dir / "all_products_extracted.json"
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(all_segments, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Données sauvegardées dans : {output_json}")
    
    # Statistiques globales
    total_products = sum(len(seg['products']) for seg in all_segments.values())
    total_brands = sum(len(seg['brands']) for seg in all_segments.values())
    
    print(f"\n📊 STATISTIQUES GLOBALES")
    print(f"  Total produits : {total_products}")
    print(f"  Total marques : {total_brands}")
    print(f"  Segments : {len(all_segments)}")
    
    return all_segments

if __name__ == "__main__":
    main()
